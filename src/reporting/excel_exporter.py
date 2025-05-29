"""
Excel Report Exporter
Export environmental assessment data to Excel format

Author: Edy Bassil
Email: bassileddy@gmail.com
"""

from typing import Dict, Any, List, Optional, Union
from datetime import datetime
import logging
import io
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.orm import Session
import pandas as pd

from src.models import (
    Project, Assessment, ImpactRecord, ComplianceRecord,
    MonitoringData, MitigationMeasure, get_session
)
from src.services import BaseService, ServiceException

logger = logging.getLogger(__name__)


class ExcelExporter(BaseService):
    """Service for exporting EIA data to Excel format."""
    
    def __init__(self, db: Session):
        super().__init__(db, Project)
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup Excel styles."""
        # Header style
        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(bold=True, color="FFFFFF", size=12)
        self.header_style.fill = PatternFill(
            start_color="1e3a5f",
            end_color="1e3a5f",
            fill_type="solid"
        )
        self.header_style.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        self.header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Subheader style
        self.subheader_style = NamedStyle(name="subheader")
        self.subheader_style.font = Font(bold=True, size=11)
        self.subheader_style.fill = PatternFill(
            start_color="2c5282",
            end_color="2c5282",
            fill_type="solid"
        )
        self.subheader_style.font.color = "FFFFFF"
        self.subheader_style.alignment = Alignment(horizontal="left", vertical="center")
        
        # Title style
        self.title_style = NamedStyle(name="title")
        self.title_style.font = Font(bold=True, size=16, color="1e3a5f")
        self.title_style.alignment = Alignment(horizontal="center", vertical="center")
        
        # Number style
        self.number_style = NamedStyle(name="number")
        self.number_style.alignment = Alignment(horizontal="right", vertical="center")
        self.number_style.number_format = '#,##0.00'
        
        # Percentage style
        self.percent_style = NamedStyle(name="percent")
        self.percent_style.alignment = Alignment(horizontal="right", vertical="center")
        self.percent_style.number_format = '0.0%'
        
        # Date style
        self.date_style = NamedStyle(name="date")
        self.date_style.alignment = Alignment(horizontal="center", vertical="center")
        self.date_style.number_format = 'yyyy-mm-dd'
        
        # Currency style
        self.currency_style = NamedStyle(name="currency")
        self.currency_style.alignment = Alignment(horizontal="right", vertical="center")
        self.currency_style.number_format = '$#,##0.00'
        
        # Status styles
        self.compliant_style = NamedStyle(name="compliant")
        self.compliant_style.fill = PatternFill(
            start_color="38a169",
            end_color="38a169",
            fill_type="solid"
        )
        self.compliant_style.font = Font(color="FFFFFF", bold=True)
        
        self.non_compliant_style = NamedStyle(name="non_compliant")
        self.non_compliant_style.fill = PatternFill(
            start_color="e53e3e",
            end_color="e53e3e",
            fill_type="solid"
        )
        self.non_compliant_style.font = Font(color="FFFFFF", bold=True)
    
    def export_project_data(
        self,
        project_id: int,
        output_path: Optional[str] = None,
        include_charts: bool = True
    ) -> Union[str, bytes]:
        """
        Export comprehensive project data to Excel.
        
        Args:
            project_id: Project ID
            output_path: Optional output file path
            include_charts: Whether to include charts
            
        Returns:
            File path if output_path provided, else Excel bytes
        """
        try:
            # Get project
            project = self.get_by_id(project_id)
            if not project:
                raise ServiceException(f"Project {project_id} not found")
            
            # Create workbook
            wb = Workbook()
            
            # Add styles to workbook
            self._add_styles_to_workbook(wb)
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets
            self._create_overview_sheet(wb, project)
            self._create_screening_sheet(wb, project)
            self._create_impact_sheet(wb, project, include_charts)
            self._create_compliance_sheet(wb, project, include_charts)
            self._create_monitoring_sheet(wb, project, include_charts)
            self._create_mitigation_sheet(wb, project)
            self._create_dashboard_sheet(wb, project, include_charts)
            
            # Save workbook
            if output_path:
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                wb.save(output_path)
                logger.info(f"Excel report saved to {output_path}")
                return output_path
            else:
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                return buffer.getvalue()
                
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise ServiceException(f"Failed to export Excel report: {str(e)}")
    
    def _add_styles_to_workbook(self, wb: Workbook):
        """Add custom styles to workbook."""
        styles = [
            self.header_style, self.subheader_style, self.title_style,
            self.number_style, self.percent_style, self.date_style,
            self.currency_style, self.compliant_style, self.non_compliant_style
        ]
        
        for style in styles:
            if style.name not in wb.named_styles:
                wb.add_named_style(style)
    
    def _create_overview_sheet(self, wb: Workbook, project: Project):
        """Create project overview sheet."""
        ws = wb.create_sheet("Project Overview")
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Environmental Impact Assessment - {project.name}"
        ws['A1'].style = self.title_style
        
        # Project details
        row = 3
        ws[f'A{row}'] = "PROJECT INFORMATION"
        ws[f'A{row}'].style = self.subheader_style
        ws.merge_cells(f'A{row}:F{row}')
        
        # Project data
        project_data = [
            ["Field", "Value"],
            ["Project Name", project.name],
            ["Project Type", project.project_type.replace('_', ' ').title()],
            ["Location", project.location],
            ["Coordinates", f"{project.latitude:.6f}, {project.longitude:.6f}" if project.latitude else "Not specified"],
            ["Status", project.status],
            ["Size", f"{project.size:,.0f} m²" if project.size else "Not specified"],
            ["Duration", f"{project.duration} months" if project.duration else "Not specified"],
            ["Budget", f"${project.budget:,.1f}M" if project.budget else "Not specified"],
            ["", ""],
            ["CLIENT INFORMATION", ""],
            ["Client Name", project.client_name or "Not specified"],
            ["Client Contact", project.client_contact or "Not specified"],
            ["Contractor", project.contractor or "Not specified"],
            ["Number of Workers", f"{project.num_workers:,}" if project.num_workers else "Not specified"],
            ["", ""],
            ["ENVIRONMENTAL PARAMETERS", ""],
            ["Water Usage", f"{project.water_usage:,.0f} m³/day" if project.water_usage else "Not specified"],
            ["Construction Area", f"{project.construction_area:,.0f} m²" if project.construction_area else "Not specified"],
            ["", ""],
            ["METADATA", ""],
            ["Created Date", project.created_at.strftime("%Y-%m-%d %H:%M")],
            ["Last Updated", project.updated_at.strftime("%Y-%m-%d %H:%M")]
        ]
        
        # Write data
        start_row = row + 2
        for i, (field, value) in enumerate(project_data):
            current_row = start_row + i
            
            if field in ["CLIENT INFORMATION", "ENVIRONMENTAL PARAMETERS", "METADATA"]:
                ws[f'A{current_row}'] = field
                ws[f'A{current_row}'].style = self.subheader_style
                ws.merge_cells(f'A{current_row}:F{current_row}')
            elif field:
                ws[f'A{current_row}'] = field
                ws[f'B{current_row}'] = value
                ws[f'A{current_row}'].font = Font(bold=True)
                ws.merge_cells(f'B{current_row}:F{current_row}')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        
        # Add description if available
        if project.description:
            desc_row = start_row + len(project_data) + 2
            ws[f'A{desc_row}'] = "PROJECT DESCRIPTION"
            ws[f'A{desc_row}'].style = self.subheader_style
            ws.merge_cells(f'A{desc_row}:F{desc_row}')
            
            ws[f'A{desc_row + 1}'] = project.description
            ws.merge_cells(f'A{desc_row + 1}:F{desc_row + 3}')
            ws[f'A{desc_row + 1}'].alignment = Alignment(wrap_text=True, vertical="top")
    
    def _create_screening_sheet(self, wb: Workbook, project: Project):
        """Create screening assessment sheet."""
        ws = wb.create_sheet("Screening Assessment")
        
        # Get latest assessment
        assessment = self.db.query(Assessment).filter_by(
            project_id=project.id
        ).order_by(Assessment.assessment_date.desc()).first()
        
        if not assessment:
            ws['A1'] = "No screening assessment data available"
            return
        
        # Title
        ws['A1'] = "SCREENING ASSESSMENT RESULTS"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:F1')
        
        # Summary
        row = 3
        ws[f'A{row}'] = "Assessment Summary"
        ws[f'A{row}'].style = self.subheader_style
        ws.merge_cells(f'A{row}:F{row}')
        
        # Results
        row += 2
        summary_data = [
            ["Parameter", "Result"],
            ["EIA Required", "YES" if assessment.eia_required else "NO"],
            ["EIA Level", assessment.eia_level or "N/A"],
            ["Assessment Type", assessment.assessment_type.title()],
            ["Assessment Date", assessment.assessment_date.strftime("%Y-%m-%d")],
            ["Estimated Duration", f"{assessment.estimated_duration} months" if assessment.estimated_duration else "N/A"],
            ["Status", assessment.status]
        ]
        
        for field, value in summary_data:
            ws[f'A{row}'] = field
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            
            # Color EIA requirement
            if field == "EIA Required":
                if value == "YES":
                    ws[f'B{row}'].style = self.non_compliant_style
                else:
                    ws[f'B{row}'].style = self.compliant_style
            
            row += 1
        
        # Key Concerns
        if assessment.key_concerns:
            row += 2
            ws[f'A{row}'] = "Key Environmental Concerns"
            ws[f'A{row}'].style = self.subheader_style
            ws.merge_cells(f'A{row}:F{row}')
            
            row += 1
            for i, concern in enumerate(assessment.key_concerns, 1):
                ws[f'A{row}'] = f"{i}."
                ws[f'B{row}'] = concern
                ws.merge_cells(f'B{row}:F{row}')
                row += 1
        
        # Regulatory Requirements
        if assessment.regulatory_requirements:
            row += 2
            ws[f'A{row}'] = "Applicable Regulatory Requirements"
            ws[f'A{row}'].style = self.subheader_style
            ws.merge_cells(f'A{row}:F{row}')
            
            row += 1
            headers = ["#", "Requirement", "Authority", "Category"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).style = self.header_style
            
            row += 1
            for i, req in enumerate(assessment.regulatory_requirements, 1):
                # Parse requirement
                parts = req.split(" - ")
                req_name = parts[0] if parts else req
                authority = parts[1] if len(parts) > 1 else "N/A"
                category = "Environmental" if "emission" in req.lower() or "quality" in req.lower() else "General"
                
                ws[f'A{row}'] = i
                ws[f'B{row}'] = req_name
                ws[f'C{row}'] = authority
                ws[f'D{row}'] = category
                row += 1
        
        # Specialist Studies
        if assessment.specialist_studies:
            row += 2
            ws[f'A{row}'] = "Required Specialist Studies"
            ws[f'A{row}'].style = self.subheader_style
            ws.merge_cells(f'A{row}:F{row}')
            
            row += 1
            for i, study in enumerate(assessment.specialist_studies, 1):
                ws[f'A{row}'] = f"{i}."
                ws[f'B{row}'] = study
                ws.merge_cells(f'B{row}:F{row}')
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
    
    def _create_impact_sheet(self, wb: Workbook, project: Project, include_charts: bool):
        """Create impact assessment sheet."""
        ws = wb.create_sheet("Impact Assessment")
        
        # Get latest impact
        impact = self.db.query(ImpactRecord).filter_by(
            project_id=project.id
        ).order_by(ImpactRecord.assessment_date.desc()).first()
        
        if not impact:
            ws['A1'] = "No impact assessment data available"
            return
        
        # Title
        ws['A1'] = "ENVIRONMENTAL IMPACT ASSESSMENT"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:G1')
        
        # Summary
        row = 3
        ws[f'A{row}'] = f"Overall Impact Severity: {impact.impact_severity}"
        severity_style = {
            'Low': self.compliant_style,
            'Medium': NamedStyle(name="medium", fill=PatternFill("solid", start_color="d69e2e")),
            'High': self.non_compliant_style
        }.get(impact.impact_severity, self.header_style)
        
        if severity_style.name not in wb.named_styles and severity_style.name == "medium":
            wb.add_named_style(severity_style)
        
        ws[f'A{row}'].style = severity_style.name if hasattr(severity_style, 'name') else self.header_style
        ws.merge_cells(f'A{row}:G{row}')
        
        # Impact Data Table
        row += 2
        ws[f'A{row}'] = "Environmental Impact Metrics"
        ws[f'A{row}'].style = self.subheader_style
        ws.merge_cells(f'A{row}:G{row}')
        
        row += 1
        headers = ["Impact Category", "Value", "Unit", "Benchmark", "Status", "% of Benchmark", "Notes"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).style = self.header_style
        
        # Impact data
        impact_data = [
            ("Carbon Footprint", impact.carbon_footprint, "tons CO₂e", 1000, 
             "Below" if impact.carbon_footprint < 1000 else "Above"),
            ("Water Consumption", impact.water_consumption, "m³", 10000,
             "Below" if impact.water_consumption < 10000 else "Above"),
            ("Waste Generation", impact.waste_generation, "tons", 100,
             "Below" if impact.waste_generation < 100 else "Above"),
            ("Energy Usage", impact.energy_usage, "MWh", 500,
             "Below" if impact.energy_usage < 500 else "Above"),
            ("Biodiversity Score", impact.biodiversity_score, "index", 70,
             "Good" if impact.biodiversity_score >= 70 else "Poor")
        ]
        
        # Add air quality if available
        if impact.pm10_concentration:
            impact_data.append(
                ("PM10 Concentration", impact.pm10_concentration, "µg/m³", 150,
                 "Below" if impact.pm10_concentration < 150 else "Above")
            )
        
        if impact.pm25_concentration:
            impact_data.append(
                ("PM2.5 Concentration", impact.pm25_concentration, "µg/m³", 65,
                 "Below" if impact.pm25_concentration < 65 else "Above")
            )
        
        row += 1
        data_start_row = row
        
        for category, value, unit, benchmark, status in impact_data:
            ws[f'A{row}'] = category
            ws[f'B{row}'] = value
            ws[f'B{row}'].style = self.number_style
            ws[f'C{row}'] = unit
            ws[f'D{row}'] = benchmark
            ws[f'E{row}'] = status
            ws[f'F{row}'] = value / benchmark if benchmark > 0 else 0
            ws[f'F{row}'].style = self.percent_style
            
            # Color status
            if status in ["Below", "Good"]:
                ws[f'E{row}'].style = self.compliant_style
            else:
                ws[f'E{row}'].style = self.non_compliant_style
            
            # Add notes
            if category == "Carbon Footprint" and value > 1000:
                ws[f'G{row}'] = "Consider carbon offset program"
            elif category == "Water Consumption" and value > 10000:
                ws[f'G{row}'] = "Implement water conservation measures"
            elif category == "Biodiversity Score" and value < 70:
                ws[f'G{row}'] = "Habitat restoration recommended"
            
            row += 1
        
        # Create chart if requested
        if include_charts:
            # Impact comparison chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Environmental Impact vs Benchmarks"
            chart.y_axis.title = "Percentage of Benchmark"
            chart.x_axis.title = "Impact Category"
            
            # Data for chart
            data = Reference(ws, min_col=6, min_row=data_start_row - 1, 
                           max_row=row - 1, max_col=6)
            cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=row - 1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            chart.width = 15
            chart.height = 10
            
            ws.add_chart(chart, f"A{row + 2}")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30
    
    def _create_compliance_sheet(self, wb: Workbook, project: Project, include_charts: bool):
        """Create compliance assessment sheet."""
        ws = wb.create_sheet("Compliance Status")
        
        # Get compliance records
        compliance_records = self.db.query(ComplianceRecord).filter_by(
            project_id=project.id
        ).all()
        
        if not compliance_records:
            ws['A1'] = "No compliance data available"
            return
        
        # Title
        ws['A1'] = "REGULATORY COMPLIANCE STATUS"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:H1')
        
        # Calculate summary
        total = len(compliance_records)
        compliant = sum(1 for r in compliance_records if r.status == "Compliant")
        non_compliant = total - compliant
        compliance_rate = (compliant / total * 100) if total > 0 else 100
        
        # Summary section
        row = 3
        ws[f'A{row}'] = f"Overall Compliance Rate: {compliance_rate:.1f}%"
        if compliance_rate >= 90:
            ws[f'A{row}'].style = self.compliant_style
        elif compliance_rate >= 70:
            ws[f'A{row}'].font = Font(color="d69e2e", bold=True)
        else:
            ws[f'A{row}'].style = self.non_compliant_style
        ws.merge_cells(f'A{row}:H{row}')
        
        # Statistics
        row += 2
        stats_data = [
            ["Metric", "Value"],
            ["Total Requirements Checked", total],
            ["Compliant Items", compliant],
            ["Non-Compliant Items", non_compliant],
            ["Compliance Rate", f"{compliance_rate:.1f}%"],
            ["Last Check Date", max(r.check_date for r in compliance_records).strftime("%Y-%m-%d")]
        ]
        
        for metric, value in stats_data:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Compliance pie chart
        if include_charts and non_compliant > 0:
            pie = PieChart()
            pie.title = "Compliance Status Distribution"
            
            # Data for pie chart
            labels = Reference(ws, min_col=1, min_row=row + 2, max_row=row + 3)
            data = Reference(ws, min_col=2, min_row=row + 2, max_row=row + 3)
            
            # Add temporary data for chart
            ws[f'A{row + 2}'] = "Compliant"
            ws[f'B{row + 2}'] = compliant
            ws[f'A{row + 3}'] = "Non-Compliant"
            ws[f'B{row + 3}'] = non_compliant
            
            pie.add_data(data)
            pie.set_categories(labels)
            pie.width = 10
            pie.height = 8
            
            ws.add_chart(pie, f"D{row - 5}")
        
        # Detailed compliance table
        row += 6
        ws[f'A{row}'] = "Detailed Compliance Checks"
        ws[f'A{row}'].style = self.subheader_style
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        headers = ["Category", "Regulation", "Status", "Actual", "Required", "Unit", "Deviation %", "Action Required"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).style = self.header_style
        
        # Sort records by category and status
        compliance_records.sort(key=lambda x: (x.category, x.status != "Compliant"))
        
        row += 1
        for record in compliance_records:
            ws[f'A{row}'] = record.category
            ws[f'B{row}'] = record.regulation_name
            ws[f'C{row}'] = record.status
            ws[f'D{row}'] = record.actual_value if record.actual_value else "N/A"
            ws[f'E{row}'] = record.required_value if record.required_value else "N/A"
            ws[f'F{row}'] = record.unit if record.unit else "N/A"
            ws[f'G{row}'] = f"{record.deviation:.1f}" if record.deviation else "0"
            ws[f'H{row}'] = record.recommendation if record.status != "Compliant" else "None"
            
            # Style status cell
            if record.status == "Compliant":
                ws[f'C{row}'].style = self.compliant_style
            else:
                ws[f'C{row}'].style = self.non_compliant_style
            
            # Style numbers
            if isinstance(record.actual_value, (int, float)):
                ws[f'D{row}'].style = self.number_style
            if isinstance(record.required_value, (int, float)):
                ws[f'E{row}'].style = self.number_style
            
            row += 1
        
        # Create table
        tab = Table(displayName="ComplianceTable", ref=f"A{row - len(compliance_records) - 1}:H{row - 1}")
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        # Adjust column widths
        column_widths = [15, 35, 12, 10, 10, 8, 12, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def _create_monitoring_sheet(self, wb: Workbook, project: Project, include_charts: bool):
        """Create monitoring data sheet."""
        ws = wb.create_sheet("Monitoring Data")
        
        # Get monitoring data (last 30 days)
        from datetime import timedelta
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)
        
        monitoring_data = self.db.query(MonitoringData).filter(
            MonitoringData.project_id == project.id,
            MonitoringData.measurement_date >= start_date
        ).order_by(MonitoringData.measurement_date.desc()).all()
        
        if not monitoring_data:
            ws['A1'] = "No monitoring data available for the last 30 days"
            return
        
        # Title
        ws['A1'] = f"ENVIRONMENTAL MONITORING DATA ({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:I1')
        
        # Summary by parameter
        row = 3
        ws[f'A{row}'] = "Monitoring Summary by Parameter"
        ws[f'A{row}'].style = self.subheader_style
        ws.merge_cells(f'A{row}:I{row}')
        
        # Group data by parameter
        from collections import defaultdict
        import numpy as np
        
        parameter_data = defaultdict(list)
        for data in monitoring_data:
            parameter_data[data.parameter].append(data)
        
        row += 1
        summary_headers = ["Parameter", "Count", "Average", "Min", "Max", "Std Dev", "Exceedances", "Compliance Rate"]
        for col, header in enumerate(summary_headers, 1):
            ws.cell(row=row, column=col, value=header).style = self.header_style
        
        row += 1
        summary_start = row
        
        for param, measurements in parameter_data.items():
            values = [m.value for m in measurements]
            exceedances = sum(1 for m in measurements if m.exceeds_limit)
            compliance_rate = ((len(measurements) - exceedances) / len(measurements) * 100) if measurements else 100
            
            ws[f'A{row}'] = param.upper()
            ws[f'B{row}'] = len(measurements)
            ws[f'C{row}'] = np.mean(values)
            ws[f'C{row}'].style = self.number_style
            ws[f'D{row}'] = min(values)
            ws[f'D{row}'].style = self.number_style
            ws[f'E{row}'] = max(values)
            ws[f'E{row}'].style = self.number_style
            ws[f'F{row}'] = np.std(values)
            ws[f'F{row}'].style = self.number_style
            ws[f'G{row}'] = exceedances
            ws[f'H{row}'] = compliance_rate / 100
            ws[f'H{row}'].style = self.percent_style
            
            # Color compliance rate
            if compliance_rate >= 95:
                ws[f'H{row}'].style = self.compliant_style
            elif compliance_rate >= 80:
                ws[f'H{row}'].font = Font(color="d69e2e", bold=True)
            else:
                ws[f'H{row}'].style = self.non_compliant_style
            
            row += 1
        
        # Time series chart
        if include_charts and len(parameter_data) > 0:
            # Create chart for first parameter with data
            first_param = list(parameter_data.keys())[0]
            param_measurements = sorted(parameter_data[first_param], key=lambda x: x.measurement_date)
            
            # Prepare data for chart
            chart_row = row + 2
            ws[f'A{chart_row}'] = "Date"
            ws[f'B{chart_row}'] = first_param.upper()
            
            for i, measurement in enumerate(param_measurements):
                ws[f'A{chart_row + i + 1}'] = measurement.measurement_date
                ws[f'A{chart_row + i + 1}'].style = self.date_style
                ws[f'B{chart_row + i + 1}'] = measurement.value
                ws[f'B{chart_row + i + 1}'].style = self.number_style
            
            # Create line chart
            chart = LineChart()
            chart.title = f"{first_param.upper()} Monitoring Trend"
            chart.style = 12
            chart.y_axis.title = f"{first_param.upper()} ({param_measurements[0].unit})"
            chart.x_axis.title = "Date"
            chart.x_axis = DateAxis()
            chart.x_axis.number_format = "dd-mmm"
            
            data = Reference(ws, min_col=2, min_row=chart_row, 
                           max_row=chart_row + len(param_measurements), max_col=2)
            dates = Reference(ws, min_col=1, min_row=chart_row + 1, 
                            max_row=chart_row + len(param_measurements))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(dates)
            chart.width = 15
            chart.height = 10
            
            ws.add_chart(chart, f"K{row - len(parameter_data) - 1}")
        
        # Detailed measurements table
        row = chart_row + len(param_measurements) + 5 if include_charts else row + 2
        ws[f'A{row}'] = "Recent Measurements (Last 100)"
        ws[f'A{row}'].style = self.subheader_style
        ws.merge_cells(f'A{row}:I{row}')
        
        row += 1
        detail_headers = ["Date", "Time", "Parameter", "Value", "Unit", "Limit", "Exceeds", "Location", "Weather"]
        for col, header in enumerate(detail_headers, 1):
            ws.cell(row=row, column=col, value=header).style = self.header_style
        
        row += 1
        for data in monitoring_data[:100]:  # Last 100 measurements
            ws[f'A{row}'] = data.measurement_date
            ws[f'A{row}'].style = self.date_style
            ws[f'B{row}'] = data.measurement_time or "N/A"
            ws[f'C{row}'] = data.parameter.upper()
            ws[f'D{row}'] = data.value
            ws[f'D{row}'].style = self.number_style
            ws[f'E{row}'] = data.unit
            ws[f'F{row}'] = data.limit_value if data.limit_value else "N/A"
            if isinstance(data.limit_value, (int, float)):
                ws[f'F{row}'].style = self.number_style
            ws[f'G{row}'] = "YES" if data.exceeds_limit else "NO"
            
            # Color exceedance
            if data.exceeds_limit:
                ws[f'G{row}'].style = self.non_compliant_style
            else:
                ws[f'G{row}'].style = self.compliant_style
            
            ws[f'H{row}'] = data.monitoring_point or "N/A"
            ws[f'I{row}'] = data.weather_conditions or "N/A"
            
            row += 1
        
        # Adjust column widths
        column_widths = [12, 10, 12, 10, 8, 10, 10, 20, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def _create_mitigation_sheet(self, wb: Workbook, project: Project):
        """Create mitigation measures sheet."""
        ws = wb.create_sheet("Mitigation Measures")
        
        # Get mitigation measures
        mitigation_measures = self.db.query(MitigationMeasure).filter_by(
            project_id=project.id
        ).all()
        
        if not mitigation_measures:
            ws['A1'] = "No mitigation measures defined"
            return
        
        # Title
        ws['A1'] = "ENVIRONMENTAL MITIGATION MEASURES"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:I1')
        
        # Summary
        row = 3
        total_measures = len(mitigation_measures)
        implemented = sum(1 for m in mitigation_measures if m.status == "Implemented")
        in_progress = sum(1 for m in mitigation_measures if m.status == "In Progress")
        planned = sum(1 for m in mitigation_measures if m.status == "Planned")
        
        ws[f'A{row}'] = f"Total Measures: {total_measures} | Implemented: {implemented} | In Progress: {in_progress} | Planned: {planned}"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:I{row}')
        
        # Measures table
        row += 2
        headers = ["Category", "Measure", "Type", "Status", "Implementation Date", 
                  "Responsible", "Cost Est.", "Expected Reduction", "Actual Reduction"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).style = self.header_style
        
        row += 1
        for measure in mitigation_measures:
            ws[f'A{row}'] = measure.impact_category
            ws[f'B{row}'] = measure.description
            ws[f'C{row}'] = measure.measure_type or "General"
            ws[f'D{row}'] = measure.status
            ws[f'E{row}'] = measure.implementation_date.strftime("%Y-%m-%d") if measure.implementation_date else "TBD"
            ws[f'F{row}'] = measure.responsible_party or "N/A"
            ws[f'G{row}'] = measure.cost_estimate if measure.cost_estimate else 0
            ws[f'G{row}'].style = self.currency_style
            ws[f'H{row}'] = f"{measure.expected_reduction}%" if measure.expected_reduction else "N/A"
            ws[f'I{row}'] = f"{measure.actual_reduction}%" if measure.actual_reduction else "N/A"
            
            # Color status
            status_colors = {
                "Implemented": self.compliant_style,
                "In Progress": NamedStyle(name="in_progress", fill=PatternFill("solid", start_color="3182ce")),
                "Planned": NamedStyle(name="planned", fill=PatternFill("solid", start_color="a0aec0"))
            }
            
            style = status_colors.get(measure.status)
            if style and style.name not in wb.named_styles and style.name in ["in_progress", "planned"]:
                wb.add_named_style(style)
            
            if style:
                ws[f'D{row}'].style = style.name if hasattr(style, 'name') else self.header_style
            
            row += 1
        
        # Cost summary
        row += 2
        total_cost = sum(m.cost_estimate for m in mitigation_measures if m.cost_estimate)
        ws[f'A{row}'] = "Total Estimated Cost:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_cost
        ws[f'B{row}'].style = self.currency_style
        
        # Adjust column widths
        column_widths = [15, 40, 12, 12, 15, 20, 12, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def _create_dashboard_sheet(self, wb: Workbook, project: Project, include_charts: bool):
        """Create executive dashboard sheet."""
        ws = wb.create_sheet("Dashboard", 0)  # Insert as first sheet
        
        # Title
        ws['A1'] = f"ENVIRONMENTAL DASHBOARD - {project.name}"
        ws['A1'].style = self.title_style
        ws['A1'].font = Font(bold=True, size=20, color="1e3a5f")
        ws.merge_cells('A1:J2')
        
        # Key Metrics Section
        row = 4
        ws[f'A{row}'] = "KEY PERFORMANCE INDICATORS"
        ws[f'A{row}'].style = self.subheader_style
        ws.merge_cells(f'A{row}:J{row}')
        
        # Calculate KPIs
        # Get latest data
        assessment = self.db.query(Assessment).filter_by(
            project_id=project.id
        ).order_by(Assessment.assessment_date.desc()).first()
        
        impact = self.db.query(ImpactRecord).filter_by(
            project_id=project.id
        ).order_by(ImpactRecord.assessment_date.desc()).first()
        
        compliance_records = self.db.query(ComplianceRecord).filter_by(
            project_id=project.id
        ).all()
        
        monitoring_data = self.db.query(MonitoringData).filter(
            MonitoringData.project_id == project.id,
            MonitoringData.measurement_date >= datetime.now() - timedelta(days=30)
        ).all()
        
        # KPI calculations
        eia_required = "YES" if assessment and assessment.eia_required else "NO"
        impact_severity = impact.impact_severity if impact else "Not Assessed"
        compliance_rate = 100
        if compliance_records:
            compliant = sum(1 for r in compliance_records if r.status == "Compliant")
            compliance_rate = (compliant / len(compliance_records) * 100)
        
        monitoring_exceedances = sum(1 for m in monitoring_data if m.exceeds_limit)
        
        # Create KPI cards
        row += 2
        kpi_data = [
            ("EIA REQUIRED", eia_required, "YES" if eia_required == "YES" else None),
            ("IMPACT LEVEL", impact_severity, "High" if impact_severity == "High" else None),
            ("COMPLIANCE", f"{compliance_rate:.1f}%", None if compliance_rate >= 90 else compliance_rate),
            ("EXCEEDANCES", str(monitoring_exceedances), None if monitoring_exceedances == 0 else monitoring_exceedances)
        ]
        
        col = 1
        for kpi_name, kpi_value, alert_condition in kpi_data:
            # KPI name
            ws.cell(row=row, column=col, value=kpi_name).font = Font(bold=True, size=10)
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            
            # KPI value
            value_cell = ws.cell(row=row + 1, column=col, value=kpi_value)
            value_cell.font = Font(bold=True, size=16)
            ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
            
            # Apply color based on condition
            if alert_condition is not None:
                value_cell.fill = PatternFill("solid", start_color="e53e3e")
                value_cell.font = Font(bold=True, size=16, color="FFFFFF")
            else:
                value_cell.fill = PatternFill("solid", start_color="38a169")
                value_cell.font = Font(bold=True, size=16, color="FFFFFF")
            
            # Alignment
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            col += 3
        
        # Project Status Section
        row += 5
        ws[f'A{row}'] = "PROJECT STATUS"
        ws[f'A{row}'].style = self.subheader_style
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        status_data = [
            ["Status", project.status],
            ["Type", project.project_type.replace('_', ' ').title()],
            ["Location", project.location],
            ["Duration", f"{project.duration} months" if project.duration else "N/A"],
            ["Progress", "Active" if project.status == "active" else "On Hold"]
        ]
        
        for field, value in status_data:
            ws[f'A{row}'] = field
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws.merge_cells(f'B{row}:E{row}')
            row += 1
        
        # Quick Actions Section
        row = 14
        ws[f'F{row}'] = "REQUIRED ACTIONS"
        ws[f'F{row}'].style = self.subheader_style
        ws.merge_cells(f'F{row}:J{row}')
        
        row += 1
        actions = []
        
        # Generate actions based on data
        if compliance_rate < 90:
            actions.append("Address regulatory compliance gaps")
        if monitoring_exceedances > 5:
            actions.append("Investigate monitoring limit exceedances")
        if impact and impact.carbon_footprint > 1000:
            actions.append("Implement carbon reduction measures")
        if not actions:
            actions.append("Continue routine monitoring")
        
        for i, action in enumerate(actions[:5], 1):
            ws[f'F{row}'] = f"{i}."
            ws[f'G{row}'] = action
            ws.merge_cells(f'G{row}:J{row}')
            row += 1
        
        # Summary Chart Section
        if include_charts and impact:
            # Environmental impact overview chart
            chart_data_row = 25
            ws[f'A{chart_data_row}'] = "Impact Metrics"
            ws[f'B{chart_data_row}'] = "% of Benchmark"
            
            impact_metrics = [
                ("Carbon", impact.carbon_footprint / 10),  # Scaled for visualization
                ("Water", impact.water_consumption / 100),
                ("Waste", impact.waste_generation),
                ("Energy", impact.energy_usage / 5),
                ("Biodiversity", 100 - impact.biodiversity_score)
            ]
            
            for i, (metric, value) in enumerate(impact_metrics):
                ws[f'A{chart_data_row + i + 1}'] = metric
                ws[f'B{chart_data_row + i + 1}'] = value
            
            # Create radar chart (using bar chart as alternative)
            chart = BarChart()
            chart.type = "bar"
            chart.style = 11
            chart.title = "Environmental Impact Overview"
            chart.y_axis.title = "Impact Metrics"
            chart.x_axis.title = "Relative Scale"
            
            data = Reference(ws, min_col=2, min_row=chart_data_row, 
                           max_row=chart_data_row + len(impact_metrics), max_col=2)
            cats = Reference(ws, min_col=1, min_row=chart_data_row + 1, 
                           max_row=chart_data_row + len(impact_metrics))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            chart.width = 12
            chart.height = 10
            
            ws.add_chart(chart, "A22")
        
        # Formatting
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['G'].width = 35
        
        # Add borders to KPI cards
        for r in range(6, 9):
            for c in [1, 4, 7, 10]:
                if c <= 10:
                    for cell in ws[f'{get_column_letter(c)}{r}:{get_column_letter(c+1)}{r}']:
                        for c in cell:
                            c.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
    
    def export_monitoring_data(
        self,
        project_id: int,
        start_date: datetime,
        end_date: datetime,
        output_path: Optional[str] = None
    ) -> Union[str, bytes]:
        """Export monitoring data to Excel for specific date range."""
        # Query monitoring data
        monitoring_data = self.db.query(MonitoringData).filter(
            MonitoringData.project_id == project_id,
            MonitoringData.measurement_date >= start_date,
            MonitoringData.measurement_date <= end_date
        ).order_by(MonitoringData.measurement_date).all()
        
        if not monitoring_data:
            raise ServiceException("No monitoring data found for specified period")
        
        # Convert to DataFrame for easier manipulation
        data_dict = []
        for record in monitoring_data:
            data_dict.append({
                'Date': record.measurement_date,
                'Time': record.measurement_time or '',
                'Parameter': record.parameter.upper(),
                'Value': record.value,
                'Unit': record.unit,
                'Limit': record.limit_value or '',
                'Exceeds Limit': 'Yes' if record.exceeds_limit else 'No',
                'Location': record.monitoring_point or '',
                'Latitude': record.latitude or '',
                'Longitude': record.longitude or '',
                'Weather': record.weather_conditions or '',
                'Equipment': record.equipment_used or '',
                'Notes': record.notes or ''
            })
        
        df = pd.DataFrame(data_dict)
        
        # Create Excel file
        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Summary sheet
                summary_df = df.groupby('Parameter').agg({
                    'Value': ['count', 'mean', 'min', 'max', 'std'],
                    'Exceeds Limit': lambda x: (x == 'Yes').sum()
                }).round(2)
                summary_df.to_excel(writer, sheet_name='Summary')
                
                # Raw data sheet
                df.to_excel(writer, sheet_name='Raw Data', index=False)
                
                # Pivot table by parameter and date
                pivot_df = df.pivot_table(
                    values='Value',
                    index='Date',
                    columns='Parameter',
                    aggfunc='mean'
                ).round(2)
                pivot_df.to_excel(writer, sheet_name='Daily Averages')
            
            return output_path
        else:
            # Return bytes
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Monitoring Data', index=False)
            
            buffer.seek(0)
            return buffer.getvalue()
    
    def export_compliance_matrix(
        self,
        project_ids: List[int],
        output_path: Optional[str] = None
    ) -> Union[str, bytes]:
        """Export compliance matrix for multiple projects."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Compliance Matrix"
        
        # Title
        ws['A1'] = "MULTI-PROJECT COMPLIANCE MATRIX"
        ws['A1'].style = self.title_style
        ws.merge_cells('A1:Z1')
        
        # Get all unique regulations
        all_regulations = set()
        project_compliance = {}
        
        for project_id in project_ids:
            project = self.get_by_id(project_id)
            if not project:
                continue
            
            compliance_records = self.db.query(ComplianceRecord).filter_by(
                project_id=project_id
            ).all()
            
            project_compliance[project.name] = {
                r.regulation_id: r.status for r in compliance_records
            }
            all_regulations.update(r.regulation_id for r in compliance_records)
        
        # Create matrix
        regulations = sorted(list(all_regulations))
        
        # Headers
        ws['A3'] = "Project / Regulation"
        ws['A3'].style = self.header_style
        
        for col, reg in enumerate(regulations, 2):
            cell = ws.cell(row=3, column=col, value=reg)
            cell.style = self.header_style
            # Rotate text for long regulation names
            cell.alignment = Alignment(text_rotation=90, horizontal="center", vertical="bottom")
        
        # Project rows
        row = 4
        for project_name, compliance in project_compliance.items():
            ws[f'A{row}'] = project_name
            ws[f'A{row}'].font = Font(bold=True)
            
            for col, reg in enumerate(regulations, 2):
                status = compliance.get(reg, "N/A")
                cell = ws.cell(row=row, column=col, value=status)
                
                if status == "Compliant":
                    cell.style = self.compliant_style
                elif status == "Non-Compliant":
                    cell.style = self.non_compliant_style
                else:
                    cell.fill = PatternFill("solid", start_color="e2e8f0")
            
            row += 1
        
        # Summary row
        row += 1
        ws[f'A{row}'] = "Compliance Rate"
        ws[f'A{row}'].font = Font(bold=True)
        
        for col, reg in enumerate(regulations, 2):
            compliant_count = sum(
                1 for proj_comp in project_compliance.values()
                if proj_comp.get(reg) == "Compliant"
            )
            total_count = sum(
                1 for proj_comp in project_compliance.values()
                if proj_comp.get(reg) in ["Compliant", "Non-Compliant"]
            )
            
            if total_count > 0:
                rate = compliant_count / total_count
                cell = ws.cell(row=row, column=col, value=rate)
                cell.style = self.percent_style
                
                if rate >= 0.9:
                    cell.fill = PatternFill("solid", start_color="38a169")
                elif rate >= 0.7:
                    cell.fill = PatternFill("solid", start_color="d69e2e")
                else:
                    cell.fill = PatternFill("solid", start_color="e53e3e")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, len(regulations) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save or return
        if output_path:
            wb.save(output_path)
            return output_path
        else:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()